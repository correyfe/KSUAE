filename='test.csv'

# import csv
# with open('test.txt', 'r') as files:
#     reader=csv.reader(files, delimiter='\t')
#     for row in reader:
#         print(row)

# import csv
# filename='test.csv'
# fields=[]
# rows=[]
# with open(filename, 'r')as csvfile:
#     csvreader=csv.reader(csvfile)
#     fields=next(csvreader)
#     for row in csvreader:
#         rows.append(row)
# print('Total rows are: %d' %(csvreader.line_num))
# print('\field names:'+ ','.join(fields))
# print('\nFirst 6 rows are:\n')
# for row in rows[:6]:
#     print(', '.join(row))

# import csv
# csv.register_dialect('myDialect', delimiter='|',
#                         skipinitialspace=True,
#                         quoting=csv.QUOTE_ALL)

# with open(filename, 'r') as file:
#     reader=csv.reader(file, dialect='myDialect')
#     for row in reader:
#         print(row)

import openpyxl

from openpyxl import Workbook
from openpyxl import load_workbook

wb = openpyxl.Workbook()

sheet = wb.active

sheet.title='List1'
sheet['A1'] = 'Hello'
sheet['A2'] = 'World'

wb.save('new_file.xlsx')

wb = load_workbook('new_file.xlsx')
sheet['B1'] = 'Hello'
sheet['B2'] = 'World'